import sys
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

import requests
import csv
import json
import os
import random
import time
import yaml
from datetime import datetime
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    print("⚠️  openpyxl not available - install with: pip install openpyxl")


# ------------------------- Configs -------------------------

# Default BigBuy to ManoMano Category Mapping (can be overridden in YAML)
# Maps BigBuy category IDs to ManoMano category IDs
# IMPORTANT: Use Level 1 categories (top-level) to avoid CATEGORY_MISSING errors
DEFAULT_CATEGORY_MAPPING = {
    19651: "20763",  # Bricolaje y herramientas (DIY & Tools) => Edilizia, materiali da costruzione (Construction materials) [Level 1]
    19656: "20204",  # Hogar y cocina (Home & Kitchen) => Arredo casa (Home furnishings) [Level 1]
    19657: "20446",  # Iluminación (Lighting) => Illuminazione (Lighting) [Level 1]
    19661: "19596",  # Jardín (Garden) => Giardino e piscine (Garden and pools) [Level 1]
    19658: "20763",  # Industria, empresas y ciencia (Industrial) => Edilizia, materiali da costruzione (Construction materials) [Level 1]
    19652: "20763",  # Coche y moto (Automotive) => Edilizia, materiali da costruzione (Construction materials) [Level 1]
    19664: "20204",  # Oficina y papelería (Office) => Arredo casa (Home furnishings - office furniture) [Level 1]
    19653: "21092",  # Electrónica (Electronics) => Elettricità (Electricity/Electronics) [Level 1]
    19756: "19596",  # Deportes y aire libre (Sports & Outdoor) => Giardino e piscine (Outdoor/Garden) [Level 1]
    19654: "20204",  # Equipaje (Luggage) => Arredo casa (Home furnishings - storage) [Level 1]
    19666: "20946",  # Productos para mascotas (Pet Products) => Animali (Animals/Pets)
}


def load_manomano_config(country_code: str) -> dict:
    """Load ManoMano configuration from YAML file"""
    script_dir = Path(__file__).parent
    config_path = script_dir / 'config' / f'manomano_{country_code.lower()}.yaml'

    if config_path.exists():
        with open(config_path, 'r') as f:
            return yaml.safe_load(f)
    else:
        print(f"Warning: Config file {config_path} not found, using defaults")
        return {}

@dataclass
class Config:
    """Configuration for feed generation"""
    margin: float = 0.43
    vat: float = 0.22
    base_price: float = 0.25
    max_price_eur: float = 1500.0
    min_price_eur: float = 10.0
    max_volume_cm3: int = 250000
    max_weight_kg: float = 45.0
    max_handling_days: int = 2  # Include stock with 0-2 day handling (can order via API)
    # ManoMano-relevant categories (no limit on products)
    manomano_categories: Dict[int, str] = field(default_factory=lambda: {
        19651: "Bricolaje y herramientas (DIY & Tools)",
        19656: "Hogar y cocina (Home & Kitchen)",
        19657: "Iluminación (Lighting)",
        19661: "Jardín (Garden)",
        19658: "Industria, empresas y ciencia (Industrial)",
        19652: "Coche y moto (Automotive)",
        19664: "Oficina y papelería (Office)",
        19653: "Electrónica (Electronics)",
        19756: "Deportes y aire libre (Sports & Outdoor)",
        19654: "Equipaje (Luggage)",
        19666: "Productos para mascotas (Pet Products)",
    })

    # BLACK FRIDAY SETTINGS - Set to True to enable, False to disable
    enable_black_friday: bool = False
    black_friday_prefix: str = "Black Friday OFFER - "

    # Shipping settings
    carrier_grid: str = "Generale"
    shipping_time: str = "5#7"

    # Category mapping (BigBuy ID -> ManoMano ID)
    category_mapping: Dict[int, str] = field(default_factory=lambda: DEFAULT_CATEGORY_MAPPING.copy())

    @classmethod
    def from_yaml(cls, country_code: str) -> 'Config':
        """Create Config from YAML file"""
        yaml_config = load_manomano_config(country_code)

        if not yaml_config:
            return cls()

        pricing = yaml_config.get('pricing', {})
        filters = yaml_config.get('filters', {})
        promotions = yaml_config.get('promotions', {})
        shipping = yaml_config.get('shipping', {})
        categories_raw = yaml_config.get('categories', {})

        # Convert category keys to int (YAML may load them as strings)
        category_mapping = {int(k): str(v) for k, v in categories_raw.items()} if categories_raw else DEFAULT_CATEGORY_MAPPING.copy()

        return cls(
            margin=pricing.get('margin', 0.43),
            vat=pricing.get('vat', 0.22),
            base_price=pricing.get('base_price', 0.25),
            max_price_eur=pricing.get('max_price_eur', 1500.0),
            min_price_eur=pricing.get('min_price_eur', 10.0),
            max_volume_cm3=filters.get('max_volume_cm3', 250000),
            max_weight_kg=filters.get('max_weight_kg', 45.0),
            max_handling_days=filters.get('max_handling_days', 2),
            enable_black_friday=promotions.get('enable_black_friday', False),
            black_friday_prefix=promotions.get('black_friday_prefix', "Black Friday OFFER - "),
            carrier_grid=shipping.get('carrier_grid', "Generale"),
            shipping_time=shipping.get('shipping_time', "5#7"),
            category_mapping=category_mapping,
        )


@dataclass
class CountryConfig:
    """Country-specific configuration"""
    locale: str
    language: str
    name: str
    currency: str
    rate: float

    @classmethod
    def get(cls, country_code: str) -> Optional['CountryConfig']:
        configs = {
            'AT': cls('de-AT', 'de', 'Austria', 'EUR', 1.0),
            'DE': cls('de-DE', 'de', 'Germany', 'EUR', 1.0),
            'IT': cls('it-IT', 'it', 'Italy', 'EUR', 1.0),
            'SK': cls('sk-SK', 'sk', 'Slovakia', 'EUR', 1.0),
            'PL': cls('pl-PL', 'pl', 'Poland', 'PLN', 4.5),
            'CZ': cls('cs-CZ', 'cs', 'Czech Republic', 'CZK', 24.0),
        }
        return configs.get(country_code.upper())


# ------------------------- API -------------------------

class BigBuyAPI:
    """Simplified BigBuy API client"""
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://api.bigbuy.eu"
        self.headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        }

    def _request(self, endpoint: str) -> Optional[List]:
        """Make API request with error handling"""
        separator = '&' if '?' in endpoint else '?'
        url = f"{self.base_url}{endpoint}{separator}t={int(time.time())}"
        try:
            r = requests.get(url, headers=self.headers, timeout=30)
            r.raise_for_status()
            return r.json()
        except requests.exceptions.RequestException as e:
            print(f"❌ API Error ({endpoint}): {e}")
            return None

    def get_category_data(self, taxonomy_id: int, language: str) -> Dict:
        """Get all data for a category in one call structure"""
        print(f"  Fetching data for category {taxonomy_id}...")
        return {
            'products': self._request(f"/rest/catalog/products.json?parentTaxonomy={taxonomy_id}") or [],
            'variations': self._request(f"/rest/catalog/productsvariations.json?parentTaxonomy={taxonomy_id}") or [],
            'stock': self._request(f"/rest/catalog/productsstockbyhandlingdays.json?parentTaxonomy={taxonomy_id}") or [],
            'var_stock': self._request(f"/rest/catalog/productsvariationsstockbyhandlingdays.json?parentTaxonomy={taxonomy_id}") or [],
            'info': self._request(f"/rest/catalog/productsinformation.json?isoCode={language}&parentTaxonomy={taxonomy_id}") or [],
            'images': self._request(f"/rest/catalog/productsimages.json?parentTaxonomy={taxonomy_id}") or []
        }


# ------------------------- Helpers -------------------------

class StockCalculator:
    """Calculate safe stock quantities"""
    @staticmethod
    def calculate_safe_quantity(bigbuy_stock: int) -> int:
        if bigbuy_stock <= 0:
            return 0
        elif bigbuy_stock <= 2:
            return 1
        elif bigbuy_stock <= 5:
            return min(2, bigbuy_stock - 1)
        elif bigbuy_stock <= 10:
            return min(5, bigbuy_stock - 2)
        elif bigbuy_stock <= 20:
            return min(10, bigbuy_stock - 3)
        elif bigbuy_stock <= 50:
            return min(25, bigbuy_stock - 5)
        else:
            return min(50, int(bigbuy_stock * 0.9))


class ProductValidator:
    """Validate products against business rules"""
    def __init__(self, config: Config, country_config: CountryConfig):
        self.config = config
        self.country = country_config
        self.stats = {
            'total': 0, 'invalid_ean': 0, 'not_new': 0,
            'no_stock': 0, 'no_info': 0, 'price_high': 0,
            'price_low': 0, 'weight_high': 0, 'volume_high': 0,
            'valid': 0, 'variants_processed': 0, 'parents_processed': 0
        }

    def validate(self, product: Dict, info: Dict, stock: int, wholesale_price: float = None, 
                 is_parent_with_variants: bool = False) -> Tuple[bool, str]:
        """Validate product or variant
        
        Args:
            is_parent_with_variants: If True, skip EAN validation (parent is just a container)
        """
        self.stats['total'] += 1

        # Use provided wholesale_price or get from product
        price_to_check = wholesale_price if wholesale_price is not None else float(product.get('wholesalePrice', 0) or 0)

        # EAN validation - skip for parents with variants (they may not have EAN)
        if not is_parent_with_variants:
            ean = str(product.get('ean13', '')).strip()
            if not ean or ean == 'None' or len(ean) != 13 or not ean.isdigit():
                self.stats['invalid_ean'] += 1
                return False, f"Invalid EAN: {ean}"

        if product.get('condition', '').upper() != 'NEW':
            self.stats['not_new'] += 1
            return False, "Not NEW condition"

        # Allow 0 stock for parent rows (variants might have stock)
        # Reject if stock is None or negative
        if stock is None or stock < 0:
            self.stats['no_stock'] += 1
            return False, "Invalid stock"

        if not info or len(info.get('name', '').strip()) < 3:
            self.stats['no_info'] += 1
            return False, "Missing product info"

        weight = float(product.get('weight', 0) or 0)
        if weight > self.config.max_weight_kg:
            self.stats['weight_high'] += 1
            return False, f"Weight too high: {weight}kg"

        volume = (float(product.get('width', 0) or 0) *
                  float(product.get('height', 0) or 0) *
                  float(product.get('depth', 0) or 0))
        if volume > self.config.max_volume_cm3:
            self.stats['volume_high'] += 1
            return False, f"Volume too high: {volume}cm³"

        price = self._calculate_price(price_to_check)
        max_price = self.config.max_price_eur * self.country.rate
        min_price = self.config.min_price_eur * self.country.rate
        if price > max_price:
            self.stats['price_high'] += 1
            return False, f"Price too high: {price}"
        if price < min_price:
            self.stats['price_low'] += 1
            return False, f"Price too low: {price}"

        self.stats['valid'] += 1
        return True, "Valid"

    def _calculate_price(self, wholesale_eur: float, delivery_cost:float = 8) -> float:
        """Calculate product price ensuring margin after VAT (delivery VAT absorbed in base_price)"""
        target_revenue = wholesale_eur * (1 + self.config.margin) + self.config.base_price + (delivery_cost * self.config.vat)
        price_eur = target_revenue / (1 - self.config.vat)
        return price_eur * self.country.rate


class DataAggregator:
    """Aggregate and organize API data"""
    @staticmethod
    def build_stock_map(stock_data: List, var_stock_data: List, max_handling_days: int = 1) -> Tuple[Dict, Dict]:
        """Build stock maps filtering by handling days

        Args:
            max_handling_days: Maximum handling days to include (default: 1 for immediate stock)
                             Set to 1: Only immediate stock (0-1 day) - matches BigBuy website
                             Set to 2: Include 1-2 day handling
                             Set to 999: Include all stock regardless of handling time
        """
        product_stock = {}
        variation_stock = {}

        for item in stock_data:
            sku = item.get('sku')
            if sku:
                # Only count stock within acceptable handling days
                total = sum(
                    s.get('quantity', 0)
                    for s in item.get('stocks', [])
                    if s.get('maxHandlingDays', 999) <= max_handling_days
                )
                product_stock[sku] = total

        for item in var_stock_data:
            sku = item.get('sku')
            if sku:
                # Only count stock within acceptable handling days
                total = sum(
                    s.get('quantity', 0)
                    for s in item.get('stocks', [])
                    if s.get('maxHandlingDays', 999) <= max_handling_days
                )
                variation_stock[sku] = total

        return product_stock, variation_stock

    @staticmethod
    def build_info_map(info_data: List) -> Dict:
        return {item['sku']: item for item in info_data}

    @staticmethod
    def build_image_map(image_data: List) -> Dict:
        result = {}
        for img_set in image_data:
            images = img_set.get('images', [])
            if images:
                result[img_set['id']] = [img.get('url', '') for img in images[:5]]
        return result

    @staticmethod
    def build_variation_map(var_data: List) -> Dict:
        """Map variations by parent product ID"""
        result = {}
        for var in var_data:
            pid = var.get('product')
            if pid:
                result.setdefault(pid, []).append(var)
        return result


# ------------------------- ManoMano generator -------------------------

class ManoManoFeedGenerator:
    """Generate rows matching ManoMano format with variant support"""
    COLS = [
        'sku', 'ean', 'sku_manufacturer', 'brand', 'mm_category', 'title', 'description',
        'image_1', 'image_2', 'image_3', 'image_4', 'image_5',
        'product_price_vat_inc', 'min_quantity', 'increment', 'quantity',
        'use_grid', 'carrier_grid_1', 'shipping_time_carrier_grid_1',
        'width', 'width_unit', 'height', 'height_unit', 'length', 'length_unit',
        'weight', 'weight_unit', 'volume', 'DisplayWeight', 'parent_sku', 'parent_title'
    ]

    def __init__(self, config: Config, country_config: CountryConfig):
        self.config = config
        self.country = country_config

    def _price_local(self, wholesale_eur: float, delivery_cost: float = 8.0) -> float:
        target_revenue = wholesale_eur * (1 + self.config.margin) + self.config.base_price + (delivery_cost * self.config.vat)
        price_eur = target_revenue / (1 - self.config.vat)
        return price_eur * self.country.rate

    def _get_with_fallback(self, variant: Dict, parent: Dict, field: str, default=0):
        """Get field from variant, fallback to parent, then default"""
        variant_value = variant.get(field)
        if variant_value is not None and variant_value != '':
            return variant_value
        parent_value = parent.get(field)
        if parent_value is not None and parent_value != '':
            return parent_value
        return default

    def create_parent_row(self, product: Dict, info: Dict, images: List, quantity: int,
                         fallback_ean: str = "", manomano_category: str = "20204") -> Dict:
        """Create parent product row (always included, even with 0 quantity)

        Args:
            fallback_ean: Use this EAN if parent has none (e.g., from first variant)
            manomano_category: ManoMano category ID for this product
        """
        price = self._price_local(float(product.get('wholesalePrice', 0) or 0))
        weight = float(product.get('weight', 0) or 0)
        length = float(product.get('depth', 0) or 0)
        width  = float(product.get('width', 0) or 0)
        height = float(product.get('height', 0) or 0)
        volume = width * height * length

        imgs = (images + ["", "", "", "", ""])[:5]

        # Handle missing parent EAN - use fallback or empty
        parent_ean = product.get('ean13')
        if not parent_ean or parent_ean == 'None' or str(parent_ean).strip() == '':
            ean_value = fallback_ean
        else:
            ean_value = str(parent_ean)

        # Apply Black Friday prefix if enabled
        base_title = (info.get('name', 'Product') or 'Product')
        if self.config.enable_black_friday:
            title = (self.config.black_friday_prefix + base_title)[:150]
        else:
            title = base_title[:150]

        return {
            'sku': str(product.get('sku', '')),
            'ean': ean_value,
            'sku_manufacturer': str(product.get('sku', '')),
            'brand': "Autres",
            'mm_category': manomano_category,
            'title': title,
            'description': (info.get('description', '') or '')[:5000],
            'image_1': imgs[0],
            'image_2': imgs[1],
            'image_3': imgs[2],
            'image_4': imgs[3],
            'image_5': imgs[4],
            'product_price_vat_inc': round(price, 2),
            'min_quantity': 1,
            'increment': 1,
            'quantity': quantity,
            'use_grid': 1,
            'carrier_grid_1': self.config.carrier_grid,
            'shipping_time_carrier_grid_1': self.config.shipping_time,
            'width': round(width, 2),
            'width_unit': 'cm',
            'height': round(height, 2),
            'height_unit': 'cm',
            'length': round(length, 2),
            'length_unit': 'cm',
            'weight': round(weight, 2),
            'weight_unit': 'kg',
            'DisplayWeight': round(weight, 2),
            'volume': round(volume, 2),
            'parent_sku': str(product.get('sku', '')),  # Parent points to itself
            'parent_title': info.get('name', '') or ''
        }

    def create_variant_row(self, variant: Dict, parent_product: Dict, parent_info: Dict,
                          images: List, quantity: int, manomano_category: str = "20204") -> Dict:
        """Create variant row with fallback to parent data

        Args:
            manomano_category: ManoMano category ID for this product (inherited from parent)
        """

        # Price: variant wholesale or parent wholesale
        variant_price = self._get_with_fallback(variant, parent_product, 'wholesalePrice', 0)
        price = self._price_local(float(variant_price))

        # Dimensions: variant or parent
        width = float(self._get_with_fallback(variant, parent_product, 'width', 0))
        height = float(self._get_with_fallback(variant, parent_product, 'height', 0))
        length = float(self._get_with_fallback(variant, parent_product, 'depth', 0))

        # Weight: variant extraWeight or parent weight
        variant_weight = variant.get('extraWeight')
        if variant_weight is not None and variant_weight != '':
            weight = float(variant_weight)
        else:
            weight = float(parent_product.get('weight', 0) or 0)

        volume = width * height * length

        # Images from parent (variants don't have separate images)
        imgs = (images + ["", "", "", "", ""])[:5]

        # Info from parent (variants inherit title/description)
        base_title = parent_info.get('name', 'Product') or 'Product'
        description = parent_info.get('description', '') or ''

        # Apply Black Friday prefix if enabled
        if self.config.enable_black_friday:
            title = (self.config.black_friday_prefix + base_title)[:150]
        else:
            title = base_title[:150]

        return {
            'sku': str(variant.get('sku', '')),
            'ean': str(variant.get('ean13', '')),
            'sku_manufacturer': str(variant.get('sku', '')),
            'brand': "Autres",
            'mm_category': manomano_category,
            'title': title,
            'description': description[:5000],
            'image_1': imgs[0],
            'image_2': imgs[1],
            'image_3': imgs[2],
            'image_4': imgs[3],
            'image_5': imgs[4],
            'product_price_vat_inc': round(price, 2),
            'min_quantity': 1,
            'increment': 1,
            'quantity': quantity,
            'use_grid': 1,
            'carrier_grid_1': self.config.carrier_grid,
            'shipping_time_carrier_grid_1': self.config.shipping_time,
            'width': round(width, 2),
            'width_unit': 'cm',
            'height': round(height, 2),
            'height_unit': 'cm',
            'length': round(length, 2),
            'length_unit': 'cm',
            'weight': round(weight, 2),
            'weight_unit': 'kg',
            'DisplayWeight': round(weight, 2),
            'volume': round(volume, 2),
            'parent_sku': str(parent_product.get('sku', '')),  # Link to parent
            'parent_title': parent_info.get('name', '') or ''
        }

    def save_csv(self, rows: List[Dict], filename: str) -> bool:
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                w = csv.DictWriter(f, fieldnames=self.COLS)
                w.writeheader()
                for r in rows:
                    w.writerow({k: r.get(k, "") for k in self.COLS})
            print(f"✅ Created {filename}")
            return True
        except Exception as e:
            print(f"❌ CSV Error: {e}")
            return False

    def save_xlsx(self, rows: List[Dict], filename: str) -> bool:
        """Save as XLSX with proper EAN formatting (no scientific notation)"""
        if not XLSX_AVAILABLE:
            print("⚠️  Skipping XLSX - openpyxl not installed")
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "ManoMano Feed"
            
            # Header row with styling
            header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, col_name in enumerate(self.COLS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
            
            # Data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, col_name in enumerate(self.COLS, 1):
                    value = row.get(col_name, "")
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # Special handling for EAN - force as text to prevent scientific notation
                    if col_name == 'ean':
                        cell.value = str(value)
                        cell.number_format = '@'  # Text format
                    else:
                        cell.value = value
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filename)
            print(f"✅ Created {filename} (EAN formatted as text)")
            return True
        except Exception as e:
            print(f"❌ XLSX Error: {e}")
            return False

    def save_json(self, data: List[Dict], stats: Dict, filename: str, csv_filename: str) -> bool:
        try:
            info = {
                "last_updated": datetime.now().isoformat(),
                "product_count": len(data),
                "validation_stats": stats,
                "stock_validation_enabled": True,
                "max_handling_days": self.config.max_handling_days,
                "stock_note": f"Only showing stock with 0-{self.config.max_handling_days} day handling (immediate availability)",
                "max_price_eur": self.config.max_price_eur,
                "min_price_eur": self.config.min_price_eur,
                "max_volume": self.config.max_volume_cm3,
                "max_weight": self.config.max_weight_kg,
                "currency": self.country.currency,
                "margin": f"{self.config.margin*100:.0f}%",
                "country": self.country.name,
                "categories_processed": list(self.config.manomano_categories.keys()),
                "feed_url": f"https://poppulseemporium.github.io/kaufland-feed/{csv_filename}"
            }
            with open(filename, 'w') as f:
                json.dump(info, f, indent=2)
            print(f"✅ Created {filename}")
            return True
        except Exception as e:
            print(f"❌ JSON Error: {e}")
            return False

    def save_html(self, data: List[Dict], stats: Dict, filename: str, csv_filename: str) -> bool:
        try:
            prices = [row['product_price_vat_inc'] for row in data]
            min_price = min(prices) if prices else 0
            max_price = max(prices) if prices else 0
            
            # Count parents vs variants
            parents = sum(1 for r in data if r['sku'] == r['parent_sku'])
            variants = len(data) - parents
            
            html = f"""<!DOCTYPE html>
<html lang="{self.country.language}">
<head>
<meta charset="UTF-8">
<title>ManoMano Feed - {self.country.name}</title>
<style>
 body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
 .header {{ background: #0ea5e9; color: white; padding: 30px; border-radius: 10px; margin-bottom: 30px; text-align: center; }}
 .stats {{ display: flex; gap: 20px; margin-bottom: 30px; flex-wrap: wrap; }}
 .stat-box {{ background: white; padding: 20px; border-radius: 10px; text-align: center; box-shadow: 0 2px 4px rgba(0,0,0,0.1); flex: 1; min-width: 150px; }}
 .stat-number {{ font-size: 28px; font-weight: bold; color: #0ea5e9; }}
 .stat-label {{ font-size: 14px; color: #666; margin-top: 5px; }}
 .feed-url {{ background: #e3f2fd; padding: 20px; border-radius: 10px; margin: 20px 0; }}
 .feed-url code {{ background: white; padding: 10px; border-radius: 5px; display: block; word-break: break-all; margin: 10px 0; }}
 table {{ border-collapse: collapse; width: 100%; background: white; border-radius: 10px; margin-top: 20px; }}
 th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
 th {{ background: #0ea5e9; color: white; }}
 .price {{ color: #16a34a; font-weight: bold; }}
 .variant {{ background: #f0f9ff; }}
 img {{ max-width: 60px; max-height: 60px; }}
</style>
</head>
<body>
 <div class="header">
   <h1>🛍️ ManoMano Product Feed</h1>
   <p><strong>Pop Pulse Emporium</strong> - {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
 </div>
 <h2>🇪🇺 {self.country.name}</h2>
 <div class="stats">
   <div class="stat-box"><div class="stat-number">{len(data):,}</div><div class="stat-label">Total Rows</div></div>
   <div class="stat-box"><div class="stat-number">{parents:,}</div><div class="stat-label">Parent Products</div></div>
   <div class="stat-box"><div class="stat-number">{variants:,}</div><div class="stat-label">Variants</div></div>
   <div class="stat-box"><div class="stat-number">{self.config.margin*100:.0f}%</div><div class="stat-label">Margin</div></div>
   <div class="stat-box"><div class="stat-number">{self.country.currency}{min_price:.2f}</div><div class="stat-label">Min Price</div></div>
   <div class="stat-box"><div class="stat-number">{self.country.currency}{max_price:.2f}</div><div class="stat-label">Max Price</div></div>
 </div>
 <div class="feed-url">
   <h3>📡 Feed URL:</h3>
   <code>https://poppulseemporium.github.io/kaufland-feed/{csv_filename}</code>
 </div>
 <h3>📊 Sample Products (First 50)</h3>
 <table>
   <tr><th>Image</th><th>SKU</th><th>Title</th><th>EAN</th><th>Price</th><th>Stock</th><th>Type</th></tr>"""
            for row in data[:50]:
                img = f'<img src="{row["image_1"]}" alt="Product">' if row.get('image_1') else 'No image'
                title = row['title'][:40] + '...' if len(row['title']) > 40 else row['title']
                row_type = "Parent" if row['sku'] == row['parent_sku'] else "Variant"
                row_class = "" if row_type == "Parent" else "variant"
                html += f"""
   <tr class="{row_class}">
     <td>{img}</td>
     <td>{row['sku']}</td>
     <td>{title}</td>
     <td>{row['ean']}</td>
     <td class="price">{self.country.currency}{row['product_price_vat_inc']:.2f}</td>
     <td><strong>{row['quantity']}</strong></td>
     <td>{row_type}</td>
   </tr>"""
            html += """
 </table>
</body>
</html>"""
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html)
            print(f"✅ Created {filename}")
            return True
        except Exception as e:
            print(f"❌ HTML Error: {e}")
            return False


# ------------------------- Main -------------------------

def main():
    print("🚀 MANOMANO FEED GENERATOR V2 (BigBuy) - WITH VARIANTS")
    print("=" * 70)

    api_key = os.getenv('BIGBUY_API_KEY')
    if not api_key:
        print("❌ BIGBUY_API_KEY environment variable not set")
        return

    country_code = os.getenv('COUNTRY_CODE', 'IT').upper()
    country_config = CountryConfig.get(country_code)
    if not country_config:
        print(f"❌ Unsupported country: {country_code}")
        return

    print(f"🌍 Country: {country_config.name}")
    print(f"💱 Currency: {country_config.currency}")

    # Load config from YAML
    config = Config.from_yaml(country_code)
    print(f"📋 Config loaded: margin={config.margin*100:.0f}%, min_price={config.min_price_eur}EUR, max_price={config.max_price_eur}EUR")

    api = BigBuyAPI(api_key)
    validator = ProductValidator(config, country_config)
    aggregator = DataAggregator()
    generator = ManoManoFeedGenerator(config, country_config)
    stock_calc = StockCalculator()

    print(f"\n📊 Processing {len(config.manomano_categories)} ManoMano-relevant categories:")
    for cat_id, cat_name in config.manomano_categories.items():
        print(f"  [{cat_id}] {cat_name}")

    all_products = []
    all_data = {'info': [], 'images': [], 'stock': [], 'var_stock': [], 'variations': []}
    # Map product ID to BigBuy category ID (for category mapping)
    product_to_category = {}

    print("\n🔄 Collecting data from categories...")
    for i, (cat_id, cat_name) in enumerate(config.manomano_categories.items(), 1):
        print(f"\n{i}/{len(config.manomano_categories)}: {cat_name} (ID: {cat_id})")
        data = api.get_category_data(cat_id, country_config.language)

        # Tag each product with its source category
        for product in data['products']:
            product_id = product.get('id')
            if product_id:
                product_to_category[product_id] = cat_id

        all_products.extend(data['products'])
        all_data['info'].extend(data['info'])
        all_data['images'].extend(data['images'])
        all_data['stock'].extend(data['stock'])
        all_data['var_stock'].extend(data['var_stock'])
        all_data['variations'].extend(data['variations'])

        print(f"  → Products: {len(data['products'])}, Variations: {len(data['variations'])}")
        time.sleep(0.5)

    print(f"\n✅ Collected {len(all_products)} products total")
    print(f"✅ Collected {len(all_data['variations'])} variations total")

    # Build lookup maps
    print(f"\n⚙️  Stock filtering: Max handling days = {config.max_handling_days} (0-{config.max_handling_days} day{'s' if config.max_handling_days > 1 else ''})")
    prod_stock, var_stock = aggregator.build_stock_map(
        all_data['stock'], all_data['var_stock'], config.max_handling_days
    )
    info_map = aggregator.build_info_map(all_data['info'])
    image_map = aggregator.build_image_map(all_data['images'])
    variation_map = aggregator.build_variation_map(all_data['variations'])

    print(f"📊 Stock entries: {len(prod_stock)} products, {len(var_stock)} variations (immediate availability only)")

    print("\n🔍 Processing products and variants...")
    rows = []
    seen_eans = set()
    
    # Debug counters
    products_with_variants = 0
    products_without_variants = 0
    variants_found = 0
    variants_added = 0

    for product in all_products:
        sku = product.get('sku')
        product_id = product.get('id')

        if not sku or sku not in info_map:
            continue

        # Get BigBuy category and map to ManoMano category
        bigbuy_cat_id = product_to_category.get(product_id)
        manomano_cat_id = config.category_mapping.get(bigbuy_cat_id, "20204")  # Default to 20204

        # Get parent stock (direct stock, not including variants)
        parent_stock = prod_stock.get(sku, 0)
        safe_parent_qty = stock_calc.calculate_safe_quantity(parent_stock)

        # Get product info and images
        info = info_map[sku]
        images = image_map.get(product_id, [])

        # Check if product has variants
        variants = variation_map.get(product_id, [])

        if variants:
            # Product HAS variants
            products_with_variants += 1
            variants_found += len(variants)
            validator.stats['parents_processed'] += 1
            
            # Validate parent (ALLOW missing EAN for parents with variants)
            parent_price = float(product.get('wholesalePrice', 0) or 0)
            is_valid, _ = validator.validate(product, info, safe_parent_qty, parent_price, 
                                            is_parent_with_variants=True)  # Skip EAN check
            
            if is_valid:
                # Get first variant's EAN for parent fallback
                first_variant_ean = ""
                if variants and variants[0].get('ean13'):
                    first_variant_ean = str(variants[0].get('ean13', ''))
                
                # Check EAN uniqueness for parent (use variant EAN if parent has none)
                parent_ean = product.get('ean13')
                if parent_ean and parent_ean != 'None':
                    parent_ean_str = str(parent_ean)
                else:
                    parent_ean_str = first_variant_ean
                
                if parent_ean_str and parent_ean_str not in seen_eans:
                    seen_eans.add(parent_ean_str)

                    # Create parent row (use first variant's EAN if parent has none)
                    parent_row = generator.create_parent_row(
                        product, info, images, safe_parent_qty,
                        fallback_ean=first_variant_ean, manomano_category=manomano_cat_id
                    )
                    rows.append(parent_row)
                
                # Process each variant
                for variant in variants:
                    variant_sku = variant.get('sku')
                    if not variant_sku:
                        continue
                    
                    # Get variant stock
                    variant_stock = var_stock.get(variant_sku, 0)
                    safe_var_qty = stock_calc.calculate_safe_quantity(variant_stock)
                    
                    # Skip variants with no stock
                    if safe_var_qty <= 0:
                        continue
                    
                    # Validate variant (using variant's own price if available)
                    variant_price = variant.get('wholesalePrice') or parent_price
                    is_var_valid, _ = validator.validate(variant, info, safe_var_qty, float(variant_price))
                    
                    if is_var_valid:
                        variant_ean = str(variant.get('ean13', ''))
                        if variant_ean not in seen_eans:
                            seen_eans.add(variant_ean)
                            validator.stats['variants_processed'] += 1
                            variants_added += 1

                            # Create variant row
                            var_row = generator.create_variant_row(
                                variant, product, info, images, safe_var_qty,
                                manomano_category=manomano_cat_id
                            )
                            rows.append(var_row)
        else:
            # Product has NO variants - treat as standalone
            products_without_variants += 1
            
            # Standalone products MUST have valid EAN
            is_valid, _ = validator.validate(product, info, safe_parent_qty, 
                                            is_parent_with_variants=False)  # Require EAN
            
            if is_valid and safe_parent_qty > 0:  # Standalone products need stock
                parent_ean = str(product.get('ean13', ''))
                if parent_ean not in seen_eans:
                    seen_eans.add(parent_ean)

                    # Create single row (parent points to itself)
                    row = generator.create_parent_row(
                        product, info, images, safe_parent_qty,
                        manomano_category=manomano_cat_id
                    )
                    rows.append(row)

        if validator.stats['total'] % 1000 == 0:
            print(f"  Processed {validator.stats['total']:,}, valid rows: {len(rows):,}")

    print(f"\n📊 Variant Processing Summary:")
    print(f"   Products with variants: {products_with_variants:,}")
    print(f"   Products without variants: {products_without_variants:,}")
    print(f"   Total variants found: {variants_found:,}")
    print(f"   Variants added to feed: {variants_added:,}")

    print(f"\n📈 Validation Results:")
    for key, value in validator.stats.items():
        if value > 0:
            print(f"   {key}: {value:,}")

    if not rows:
        print("❌ No valid products found")
        return

    print(f"\n✅ {len(rows)} total rows ready (parents + variants)")

    # Debug: Show some examples of parent-child relationships
    print("\n🔍 Sample parent-child relationships:")
    parent_skus = {}
    for row in rows:
        ps = row['parent_sku']
        parent_skus.setdefault(ps, []).append(row['sku'])
    
    # Show families with multiple members
    families = [(ps, skus) for ps, skus in parent_skus.items() if len(skus) > 1]
    if families:
        print(f"  Found {len(families)} product families with variants:")
        for ps, skus in families[:5]:  # Show first 5
            print(f"    Parent {ps} has {len(skus)} members: {', '.join(skus[:3])}{'...' if len(skus) > 3 else ''}")
    else:
        print("  ⚠️  No product families found (all standalone products)")

    # Always use _cc suffix
    cc = country_code.lower()
    csv_file  = f'manomano_feed_{cc}.csv'
    xlsx_file = f'manomano_feed_{cc}.xlsx'
    html_file = f'manomano_index_{cc}.html'
    json_file = f'manomano_info_{cc}.json'

    print("\n📁 Creating files...")
    generator.save_csv(rows, csv_file)
    generator.save_xlsx(rows, xlsx_file)  # XLSX with proper EAN formatting
    generator.save_json(rows, validator.stats, json_file, csv_file)
    generator.save_html(rows, validator.stats, html_file, csv_file)

    print("\n" + "=" * 70)
    print("🎉 FEED GENERATION COMPLETE")
    print("=" * 70)
    print(f"📊 Total rows: {len(rows)}")
    print(f"👪 Parent products: {validator.stats['parents_processed']}")
    print(f"🔀 Variants: {validator.stats['variants_processed']}")
    print(f"✅ All products from {len(config.manomano_categories)} ManoMano categories")
    print("\n✓ FEATURES:")
    print("  • Variants processed with parent fallback")
    print("  • Parent rows always included (even with 0 stock)")
    print("  • Category whitelist (11 ManoMano-relevant categories)")
    print("  • No product limit (all valid products included)")


if __name__ == "__main__":
    main()
    
## TO RUN: 
# $env:BIGBUY_API_KEY="YjEzYWU2YTRkNmQyZTY1MjU5M2IzYjlmN2Q2OTQyMTljMjIxZjE0MTdkZGE1NTRjY2YzMTg3OWExYjllNTUzZQ"
# $env:COUNTRY_CODE="IT"
# python manomano_feed_generator.py





